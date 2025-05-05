"""
Utility functions for the inventory app
"""
import io
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from django.utils import timezone
from django.db.models import F

def generate_inventory_pdf(products):
    """
    Generate a PDF report of inventory products
    
    Args:
        products: QuerySet of Product objects
    
    Returns:
        PDF file as bytes
    """
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter)
    elements = []
    
    # Define styles
    styles = getSampleStyleSheet()
    title_style = styles['Heading1']
    subtitle_style = styles['Heading2']
    normal_style = styles['Normal']
    
    # Add title
    elements.append(Paragraph('Inventory Report', title_style))
    elements.append(Paragraph(f'Generated on: {timezone.now().strftime("%Y-%m-%d %H:%M")}', normal_style))
    elements.append(Spacer(1, 12))
    
    # Create table data
    data = [['SKU', 'Product Name', 'Category', 'Supplier', 'Quantity', 'Price', 'Value', 'Status']]
    
    # Add product data
    total_value = 0
    for product in products:
        value = product.price * product.quantity
        total_value += value
        
        status = 'Low Stock' if product.quantity < product.threshold_quantity else 'In Stock'
        
        data.append([
            product.sku,
            product.name,
            product.category.name if product.category else 'N/A',
            product.supplier.name if product.supplier else 'N/A',
            product.quantity,
            f"${product.price:.2f}",
            f"${value:.2f}",
            status
        ])
    
    # Add total row
    data.append(['', '', '', '', '', 'Total Value:', f"${total_value:.2f}", ''])
    
    # Create table
    table = Table(data, colWidths=[70, 120, 80, 80, 50, 60, 60, 60])
    
    # Style the table
    table_style = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.lightblue),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, -1), (-1, -1), colors.lightgrey),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('ALIGN', (4, 1), (7, -2), 'CENTER'),
        ('ALIGN', (5, -1), (6, -1), 'RIGHT'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ])
    
    # Add conditional formatting for low stock
    for i, product in enumerate(products, 1):
        if product.quantity < product.threshold_quantity:
            table_style.add('TEXTCOLOR', (7, i), (7, i), colors.red)
    
    table.setStyle(table_style)
    elements.append(table)
    
    # Build the PDF
    doc.build(elements)
    pdf = buffer.getvalue()
    buffer.close()
    
    return pdf

def generate_inventory_excel(products):
    """
    Generate an Excel report of inventory products
    
    Args:
        products: QuerySet of Product objects
    
    Returns:
        Excel file as bytes
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Inventory Report"
    
    # Define headers
    headers = ['SKU', 'Product Name', 'Category', 'Supplier', 'Quantity', 
               'Threshold', 'Price ($)', 'Value ($)', 'Status']
    
    # Apply header formatting
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Write headers
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Write data
    low_stock_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
    
    row_num = 2
    total_value = 0
    
    for product in products:
        value = product.price * product.quantity
        total_value += value
        
        status = 'Low Stock' if product.quantity < product.threshold_quantity else 'In Stock'
        
        row = [
            product.sku,
            product.name,
            product.category.name if product.category else 'N/A',
            product.supplier.name if product.supplier else 'N/A',
            product.quantity,
            product.threshold_quantity,
            float(product.price),
            float(value),
            status
        ]
        
        for col_num, cell_value in enumerate(row, 1):
            cell = ws.cell(row=row_num, column=col_num, value=cell_value)
            
            # Apply special formatting
            if col_num in [5, 6, 7, 8]:  # Quantity, Price, Value columns
                cell.alignment = Alignment(horizontal="center")
            
            # Highlight low stock
            if col_num == 9 and status == 'Low Stock':
                cell.fill = low_stock_fill
                cell.font = Font(color="CC0000", bold=True)
        
        row_num += 1
    
    # Add total row
    total_row = ['', 'TOTAL', '', '', '', '', '', float(total_value), '']
    for col_num, cell_value in enumerate(total_row, 1):
        cell = ws.cell(row=row_num, column=col_num, value=cell_value)
        cell.font = Font(bold=True)
        
        if col_num == 8:
            cell.alignment = Alignment(horizontal="center")
    
    # Format the sheet
    for column_cells in ws.columns:
        length = max(len(str(cell.value or "")) for cell in column_cells)
        ws.column_dimensions[column_cells[0].column_letter].width = length + 5
    
    # Add report metadata
    ws.merge_cells('A{}:I{}'.format(row_num + 2, row_num + 2))
    report_date = ws.cell(row=row_num + 2, column=1, 
                          value=f"Report generated on: {timezone.now().strftime('%Y-%m-%d %H:%M')}")
    report_date.alignment = Alignment(horizontal="left")
    
    # Convert to file
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output.getvalue()
