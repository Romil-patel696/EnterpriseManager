"""
Utility functions for the attendance app
"""
import io
from reportlab.lib.pagesizes import letter, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from django.utils import timezone
from datetime import datetime, timedelta

def generate_attendance_report_pdf(attendance_records, start_date, end_date):
    """
    Generate a PDF report of attendance records
    
    Args:
        attendance_records: QuerySet of Attendance objects
        start_date: Start date for the report period
        end_date: End date for the report period
    
    Returns:
        PDF file as bytes
    """
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(letter))
    elements = []
    
    # Define styles
    styles = getSampleStyleSheet()
    title_style = styles['Heading1']
    subtitle_style = styles['Heading2']
    normal_style = styles['Normal']
    
    # Add title
    elements.append(Paragraph('Attendance Report', title_style))
    elements.append(Paragraph(f'Period: {start_date} to {end_date}', subtitle_style))
    elements.append(Paragraph(f'Generated on: {timezone.now().strftime("%Y-%m-%d %H:%M")}', normal_style))
    elements.append(Spacer(1, 12))
    
    # Create table data
    data = [['Employee', 'Date', 'Check In', 'Check Out', 'Hours', 'Status', 'Notes']]
    
    # Group by employee
    employee_records = {}
    for record in attendance_records:
        employee_name = record.employee.get_full_name() or record.employee.username
        if employee_name not in employee_records:
            employee_records[employee_name] = []
        employee_records[employee_name].append(record)
    
    # Add data for each employee
    for employee_name, records in employee_records.items():
        # Add employee header
        data.append([employee_name, '', '', '', '', '', ''])
        
        # Add records for this employee
        for record in records:
            duration = record.get_duration() if record.check_out else '-'
            if duration:
                duration = f"{duration:.2f} hrs"
            
            check_in_time = record.check_in.strftime('%H:%M') if record.check_in else '-'
            check_out_time = record.check_out.strftime('%H:%M') if record.check_out else '-'
            
            data.append([
                '',  # No need to repeat employee name
                record.date.strftime('%Y-%m-%d'),
                check_in_time,
                check_out_time,
                duration,
                record.get_status_display(),
                record.notes or ''
            ])
        
        # Add empty row after each employee
        data.append(['', '', '', '', '', '', ''])
    
    # Create table
    col_widths = [120, 80, 70, 70, 70, 80, 200]
    table = Table(data, colWidths=col_widths)
    
    # Style the table
    style = TableStyle([
        # Header row
        ('BACKGROUND', (0, 0), (-1, 0), colors.lightblue),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        
        # Grid
        ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
        
        # Align center for certain columns
        ('ALIGN', (1, 0), (5, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ])
    
    # Apply employee header styling
    for i, row in enumerate(data):
        if row[0] and not row[1]:  # This is an employee header row
            style.add('BACKGROUND', (0, i), (-1, i), colors.lightgrey)
            style.add('FONTNAME', (0, i), (0, i), 'Helvetica-Bold')
            style.add('SPAN', (0, i), (-1, i))  # Span all columns
            style.add('ALIGN', (0, i), (0, i), 'LEFT')
    
    table.setStyle(style)
    elements.append(table)
    
    # Build the PDF
    doc.build(elements)
    pdf = buffer.getvalue()
    buffer.close()
    
    return pdf

def generate_attendance_report_excel(attendance_records, start_date, end_date):
    """
    Generate an Excel report of attendance records
    
    Args:
        attendance_records: QuerySet of Attendance objects
        start_date: Start date for the report period
        end_date: End date for the report period
    
    Returns:
        Excel file as bytes
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Attendance Report"
    
    # Define styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    employee_font = Font(bold=True)
    employee_fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Add report title and metadata
    ws.merge_cells('A1:G1')
    title_cell = ws.cell(row=1, column=1, value=f"Attendance Report: {start_date} to {end_date}")
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = Alignment(horizontal="center")
    
    ws.merge_cells('A2:G2')
    subtitle_cell = ws.cell(row=2, column=2, value=f"Generated on: {timezone.now().strftime('%Y-%m-%d %H:%M')}")
    subtitle_cell.alignment = Alignment(horizontal="center")
    
    # Add headers (row 4)
    headers = ['Employee', 'Date', 'Check In', 'Check Out', 'Hours', 'Status', 'Notes']
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Add data
    row_num = 5
    current_employee = None
    
    for record in attendance_records:
        employee_name = record.employee.get_full_name() or record.employee.username
        
        # If this is a new employee, add a header row
        if employee_name != current_employee:
            current_employee = employee_name
            
            # Skip a row between employees (except for first employee)
            if row_num > 5:
                row_num += 1
            
            # Add employee header
            for col_num in range(1, 8):
                cell = ws.cell(row=row_num, column=col_num, value=employee_name if col_num == 1 else "")
                cell.font = employee_font
                cell.fill = employee_fill
                cell.border = thin_border
            
            # Merge the employee header row
            ws.merge_cells(f'A{row_num}:G{row_num}')
            row_num += 1
        
        # Calculate hours worked
        duration = record.get_duration() if record.check_out else None
        if duration:
            duration = round(duration, 2)
        
        # Format times
        check_in_time = record.check_in.strftime('%H:%M') if record.check_in else '-'
        check_out_time = record.check_out.strftime('%H:%M') if record.check_out else '-'
        
        # Add record data
        data = [
            '',  # No need to repeat employee name
            record.date,
            check_in_time,
            check_out_time,
            duration,
            record.get_status_display(),
            record.notes or ''
        ]
        
        for col_num, value in enumerate(data, 1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            cell.border = thin_border
            
            # Center align certain columns
            if 2 <= col_num <= 6:
                cell.alignment = Alignment(horizontal="center")
        
        row_num += 1
    
    # Set column widths
    column_widths = [25, 15, 12, 12, 12, 15, 40]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[chr(64 + i)].width = width
    
    # Convert to file
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output.getvalue()

def calculate_leave_balance(employee, year=None):
    """
    Calculate leave balance for an employee
    
    Args:
        employee: User object
        year: Year to calculate balance for (defaults to current year)
    
    Returns:
        Dictionary with leave balances
    """
    from .models import LeaveBalance, Leave, LeaveType
    
    year = year or timezone.now().year
    
    # Get or create leave balance
    try:
        balance = LeaveBalance.objects.get(employee=employee, year=year)
    except LeaveBalance.DoesNotExist:
        # Get default values from policy or use hardcoded defaults
        from .views import get_policy_value
        casual_leave_default = get_policy_value('default_casual_leaves', 12)
        sick_leave_default = get_policy_value('default_sick_leaves', 10)
        
        balance = LeaveBalance.objects.create(
            employee=employee,
            year=year,
            casual_leave=casual_leave_default,
            sick_leave=sick_leave_default
        )
    
    # Get approved leaves for the year
    approved_leaves = Leave.objects.filter(
        employee=employee,
        status=Leave.APPROVED,
        start_date__year=year
    ).select_related('leave_type')
    
    # Calculate used leave days
    casual_days_used = sum([
        leave.get_business_days() 
        for leave in approved_leaves 
        if leave.leave_type.name.lower() == 'casual leave'
    ])
    
    sick_days_used = sum([
        leave.get_business_days() 
        for leave in approved_leaves 
        if leave.leave_type.name.lower() == 'sick leave'
    ])
    
    return {
        'casual': {
            'total': balance.casual_leave + casual_days_used,
            'used': casual_days_used,
            'remaining': balance.casual_leave
        },
        'sick': {
            'total': balance.sick_leave + sick_days_used,
            'used': sick_days_used,
            'remaining': balance.sick_leave
        }
    }

def mark_absent_employees():
    """
    Mark employees as absent if they haven't checked in today
    This would be run as a scheduled task at the end of each day
    """
    from .models import Attendance
    from accounts.models import User, UserProfile
    from django.db.models import Q
    
    today = timezone.now().date()
    
    # Get all employees
    employees = User.objects.filter(userprofile__role=UserProfile.EMPLOYEE)
    
    # Find employees who haven't checked in today
    for employee in employees:
        if not Attendance.objects.filter(employee=employee, date=today).exists():
            # Create an "absent" record
            Attendance.objects.create(
                employee=employee,
                date=today,
                check_in=None,
                check_out=None,
                status='absent',
                notes='Automatically marked as absent'
            )
